from openpyxl import Workbook,load_workbook
import openpyxl.utils
from openpyxl.styles import Border,Side,PatternFill
import os
import datetime
import PySimpleGUI as sg
import numpy as np
import pickle
import matplotlib.pyplot as plt
import matplotlib.animation as ani

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openxlhelper
os.chdir('C:/Users/Chua Wei Yang/Desktop/Project/Ouroborus/Data/Finance')
wb = load_workbook('FINANCE.xlsx')
# grab the active worksheet
ws = wb.active



b_order = Border(left=Side(border_style = 'thin',color='00000000'),right=Side(border_style = 'thin',color='00000000'),top=Side(border_style = 'thin',color='00000000'),bottom=Side(border_style = 'thin',color='00000000'))
# DEFAULT SAVE 

class SETTING_SAVE():
    def __init__(self):
        file = open('TEST.dat','wb')
        self.account = 'TEST'
        self.rows= ['INCOME','FOOD','TRANSPORT','FUN','OTHERS']
        self.default_values = {'INCOME':0.0,'FOOD':0.0,'FUN':0.0,'OTHERS':0.0,'TRANSPORT':3.2}
        self.daily_update = False
        self.income_per_month=0.0
        self.Goal = 0.30
        self.month = 'Feb-2021'
        self.auto_Values = {'INCOME':0.0,'FOOD':0.0,'FUN':0.0,'OTHERS':0.0,'TRANSPORT':3.2}
        self.negative = {'INCOME':False,'FOOD':True,'FUN':True,'OTHERS':False,'TRANSPORT':True}
    def save(self):
        os.chdir('C:/Users/Chua Wei Yang/Desktop/Project/Ouroborus/Data/Finance')
        pickle.dump(self,open(self.account+'.dat','wb'))

    def load(self,file):
        os.chdir('C:/Users/Chua Wei Yang/Desktop/Project/Ouroborus/Data/Finance')
        self.__dict__.update(pickle.load(open(file+'.dat','rb')).__dict__)

             
    def add_row(self,catagory,state):
        self.row.append(catagory)
        self.default_values[catagory] = 0.0
        self.auto_Values[catagory] = 0.0
        if state:
            self.negative[catagory]=True
        else:
            self.negative[catagory] = False
        
    def change_default_value (self,catagory,new_value,state):
        self.default_values[catagory] = new_value
        self.negative[catagory] = state

#----------------------------Graph---------------
def draw_pie(labels,w):
    #---------extract data-----
    row = ws['AG4':'AG7']
    data=[]
    labels.append('SAVED')
    for value in row: 
        
        data.append(abs(openxlhelper.openxl_helper(ws,value[0].value)))
        
    data.append(openxlhelper.openxl_helper(ws,ws['H15'].value))

    #---------draw pie-----
    fg,ax =plt.subplots()
    fg.set_size_inches(5, 5)
    wedges,text,pct=ax.pie(data,autopct='%.2f%%',wedgeprops=dict(width=0.5),pctdistance =0.8)
    arrowprop = {'arrowstyle':'-','connectionstyle':"angle,angleA={A:},angleB={B:},rad=0"}
    box = {'boxstyle':'round'}
    
     #get middle of wedges
    for index,part in enumerate(wedges):
        ang = ((part.theta2-part.theta1)/2) +part.theta1
        x = np.cos(ang*np.pi/180)
        y = np.sin(ang*np.pi/180)
        if ang>=0 and ang<=90: 
            arrowprop['connectionstyle'] = arrowprop['connectionstyle'].format(A=0,B=90)
            end_x = x+0.3
            end_y= y+0.3
        elif ang>90 and ang <= 180:
            arrowprop['connectionstyle'] = arrowprop['connectionstyle'].format(A=0,B=90)
            end_x = x-0.6
            end_y= y+0.3
        elif ang >180 and ang <=270:
            arrowprop['connectionstyle'] = arrowprop['connectionstyle'].format(A=0,B=90)
            end_x = x-0.6
            end_y= y-0.3
        else:
            arrowprop['connectionstyle'] = arrowprop['connectionstyle'].format(A=0,B=90)
            end_x = x+0.3
            end_y= y-0.3

        ax.annotate(labels[index],(x,y),arrowprops=arrowprop,xytext = (end_x,end_y),bbox=box)
    ax.set_title('Expenditure')
    legend = []
    for value in range(len(data)):

        legend.append('$'+ '%.2f' % data[value] +' | '+ labels[value] )
   
    
    ax.legend(loc=0,bbox_to_anchor=(-0.4, 0.25, 0.6, 0),labels=legend)

    #---------display pie-----
    fg.patch.set_facecolor(sg.theme_background_color())

    figure_canvas_agg = FigureCanvasTkAgg(fg, w['pie'].TKCanvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    #animator = ani.FuncAnimation(fg, update, interval = 100)
def sort(data):
  

    for i in range(len(data)-1):
        for j in range(0,len(data) -i-1): 
            if data[j][1] < data[j+1][1] : 
                data[j], data[j+1] = data[j+1], data[j] 
    return data
def helper(ws,data):
    one_d = []
    for i in data :
        
        one_d.append(openxlhelper.openxl_helper(ws,i[0].value))
    return one_d
def bar_graph(lable,w):
    # extract data 
    
    x = np.arange(len(lable))
    width =0.35
    this_month = datetime.datetime.now()
    last_month = this_month - datetime.timedelta(weeks = 4)
    this_month_ws = wb[this_month.strftime('%h-%Y') ]
    last_month_ws = wb[last_month.strftime('%h-%Y') ]
    this_month_data = this_month_ws['AG3':'AG8']
    last_month_data = last_month_ws['AG3':'AG8']
    # draw graph 
    fg,ax =plt.subplots()

    fg.set_size_inches(5, 5)
    ax.bar(x - width/2,helper(this_month_ws,this_month_data),width = width,label= this_month.strftime('%h-%Y'))
    ax.bar(x + width/2,helper(last_month_ws,last_month_data),width = width,label=last_month.strftime('%h-%Y'))
    ax.set_xticks(x)
    ax.set_xticklabels(helper(ws,lable))
    ax.legend()
    plt.axhline(y=0,color='black')
    fg.patch.set_facecolor(sg.theme_background_color())
    fg.tight_layout()


    figure_canvas_agg = FigureCanvasTkAgg(fg, w['bar'].TKCanvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
def print_statments():

    this_month = datetime.datetime.now()
    last_month = this_month - datetime.timedelta(weeks = 4)
    this_month_ws = wb[this_month.strftime('%h-%Y') ]
    last_month_ws = wb[last_month.strftime('%h-%Y') ]
    last_month_data = helper(last_month_ws,last_month_ws['AG4':'AG7'])
    this_month_data = helper(this_month_ws,last_month_ws['AG4':'AG7'])
    this_month_data.append(openxlhelper.openxl_helper(this_month_ws,ws['H16'].value))
    catagory =helper(this_month_ws,last_month_ws['A4':'A7'])
    percentage=[]
    if this_month_data[-1] < save_file.Goal:
        first_line =  """This month saving have dibbed below """ + str(save_file.Goal*100) + """%, current at """  + str(this_month_data[-1]*100) +'%.\n'
    else: 
        first_line = """You still got """+ str((abs(this_month_data[-1]-save_file.Goal))*100)+"% left before your goal, current at" + str(this_month_data[-1]*100) +'%.\n'
    for i in range(len(this_month_data[:-1])):
        percentage.append((round(this_month_data[i]/last_month_data[i],2)*100))
    second_line = 'The top 5 spendings:\n'
    lst =[]
    for i in range(len(percentage)):
        lst.append((str(catagory[i]),percentage[i]) )      
    lst = sort(lst)
    for i in lst : 
        second_line += i[0] + str(i[1]) +'%\n'
    third_line ='I suggest cutting back on '+ lst[0][0] +'\n'
        


         
    return first_line+  second_line + third_line
#------------------------------------start up----------------------------------
def check_month():
    global ws,wb
    this_month = datetime.datetime.now().strftime('%h-%Y')
    if this_month not in wb.sheetnames:
        create_sheet(wb,this_month)
        ws = wb[this_month]

#------------------------------------------------------------------------------
# ----------------------------CREATE FUNCTIONS--------------------------------
green = PatternFill( fill_type='solid',start_color='C6E0B4')
orange = PatternFill( fill_type='solid',start_color='F4B084')

def create_sheet(wb,month):

    ws =wb.create_sheet(month)
    table1 = ws['A2':'AG8']
    table2 = ws['G14':'H18']
    create_table(ws,b_order,table_range = table1,Left_colour=green,top_colour=green)
    create_table (ws,b_order,table_range=table2,Left_colour= orange)

    write(ws,['DATE','INCOME','FOOD','TRANSPORT','FUN','OTHERS','TOTAL'],ws['A2':'A8'])
    write(ws,[i+1 for i in range(31)],ws['B2':'AF2'],col=False)
    write(ws,['TOTAL INCOME','INCOME SAVED','% INCOME SAVED','GOAL','=IF(H18>0,"AMT LEFT","AMT NEEDED")'],ws['G14':'G18'])
    write(ws,['Sum','=SUM(B3:AF3)','=SUM(B4:AF4)','=SUM(B5:AF5)','=SUM(B6:AF6)','=SUM(B7:AF7)','=SUM(AG3:AG7)'],ws['AG2':'AG8'])
    write(ws,['=SUM('+openpyxl.utils.cell.get_column_letter(2+i)+'3:'+openpyxl.utils.cell.get_column_letter(2+i)+'7)' for i in range(31)],ws['B8':'AF8'],col=False)
    write(ws,['=AG3','=AG8','=H15/H14','30%','=(H16-H17)*H14'],ws['H14':'H18'])
    save()
    
def write(ws,texts,table_range,col=True):
    
    if col:
        for index, cell in enumerate(table_range):
            cell[0].value = texts[index]
    else:
        for index, cell in enumerate(table_range[0]):
            cell.value = texts[index]
def create_table(ws,border_style,table_range,top_colour = None,Left_colour= None,Bottom_colour = None,Right_colour = None,table_colour=None):
    f_row = table_range[0]
    l_row = table_range[len(table_range)-1]
    for row in table_range:

        if Left_colour!= None:
            row[0].fill = Left_colour
        if Right_colour!= None:
            row[len(row)-1].fill = Right_colour
        
        for col in row:

            col.border = border_style   
            if top_colour != None :
                if col in f_row:
                    col.fill = top_colour
            if Bottom_colour !=None:
                if col in l_row :
                    col.fill = Bottom_colour

#-----------------------------------------------------------------------------

def add_row():
    ws.insert_rows(15)
    row = ws['C15':'E15']
    for i in row[0]:
        i.border = b_order


def save():
    # Save the file
    wb.save("FINANCE.xlsx")




def Update(date,data):
    check_month()
    calander = ws['B2':'AF2']
    y=calander[0][int(date)-1].column
    col = openpyxl.utils.get_column_letter(y)
    for i in range(len(data)):
        row = 3+ i 
        ws[col+str(row)] = data[i]
    save()
#--------------------------------setting-------------------------------
def save_setting(file):

    file.save()
    print('SAVED')






#--------------------------------------------------MAIN----------------------------------------------------------    
#SAVE 
save_file = SETTING_SAVE()

#STATAS



Table = [[sg.Text(row[0].value +' : '+ str(openxlhelper.openxl_helper(ws,row[1].value)))] for row in ws['G14':'H18']]
Dialogue = [[sg.Text(print_statments())]]
Stats = [[sg.Frame('Chart',[[sg.Canvas(size=(10,10),key='pie',background_color='blue')]]),sg.Frame('BAr',[[sg.Canvas(key='bar',background_color='blue',size=(10,10))]])],
            [sg.Frame('Info',Table),sg.Frame('Summary',Dialogue)]]


#UPDATE TAB
update =[[sg.Text('DATE: '+datetime.datetime.today().strftime('%Y-%m-%d'))],
[sg.Text('INCOME',size=(10,1)),sg.Input(save_file.default_values['INCOME'],(10,20),key='income',metadata='+')],
[sg.Text('FOOD',size=(10,1)),sg.Input(save_file.default_values['FOOD'],(10,20),key=1,metadata='-')],
[sg.Text('TRANSPORT',size=(10,1)),sg.Input(save_file.default_values['TRANSPORT'],(10,20),key=2,metadata='-')],
[sg.Text('FUN',size=(10,1)),sg.Input(save_file.default_values['FUN'],(10,20),key=3,metadata='-')],
[sg.Text('OTHERS',size=(10,1)),sg.Input(save_file.default_values['OTHERS'],(10,20),key='others',metadata='+')],
[sg.Button('Update',key = Update)]
]
#edit TAB   
edit = [[sg.Text('TYPE',size=(10,1)),sg.DropDown(values=['INCOME','FOOD','TRANSPORT','FUN','OTHERS'],size=(10,1))],[sg.CalendarButton('Date',size=(9,1),format="%Y-%m-%d",target=(1,1)),sg.Input(default_text=datetime.datetime.today().strftime('%Y-%m-%d'),size=(10,20))],[sg.Text('$',size=(10,1)),sg.Input(size=(10,20))]]

#setting TAB
setting=[[sg.Text('Expected Income Per Month'),sg.Input((0.0),(10,20))],[sg.Text('Auto update'),sg.Combo(['ON','OFF'],key = 'auto',default_value='OFF',size=(5,1))]]
#auto-update setting
auto = [
[sg.Text('INCOME',size=(10,1)),sg.Combo(['+','-'],default_value='+',size = (1,1),font='Courier 13'),sg.Input(save_file.default_values['INCOME'],(10,20),key='INCOME')],
[sg.Text('FOOD',size=(10,1)),sg.Combo(['+','-'],default_value='+',size = (1,1),font='Courier 13'),sg.Input(save_file.default_values['FOOD'],(10,20),key='FOOD')],
[sg.Text('TRANSPORT',size=(10,1)),sg.Combo(['+','-'],default_value='+',size = (1,1),font='Courier 13'),sg.Input(save_file.default_values['TRANSPORT'],(10,20),key='TRANSPORT')],
[sg.Text('FUN',size=(10,1)),sg.Combo(['+','-'],default_value='+',size = (1,1),font='Courier 13'),sg.Input(save_file.default_values['FUN'],(10,20),key='FUN')],
[sg.Text('OTHERS',size=(10,1)),sg.Combo(['+','-'],default_value='+',size = (1,1),font='Courier 13'),sg.Input(save_file.default_values['OTHERS'],(10,20),key='OTHERS')]]
#MAin

main =[[sg.TabGroup( [ [sg.Tab('stats',Stats),sg.Tab('Update',layout = update),sg.Tab('Edit',edit),sg.Tab('Setting',layout=[[sg.Frame('Setting',setting),sg.Frame('Auto',auto)],[sg.Button('SAVE',key = save_setting)]])]])]]
window = sg.Window('Window Title', main,finalize=True)

#draw pie
draw_pie(save_file.rows[1:],window)     
bar_graph(ws['A3':'A8'],window)



# -----------------------Window--------------------------
while True:
    event, values = window.read()
    print(event)
    if event == sg.WINDOW_CLOSED or event == 'Quit':
        break
    if event ==Update:
        temp=[float(window['income'].get())]
        for i in range (3):
            if window[i+1].metadata == '-':
                temp.append(-1*float(window[i+1].get()))
            else:
                temp.append(float(window[i+1].get()))
        temp.append(float(window['others'].get()))
        Update(datetime.datetime.today().strftime('%d'),temp)
    if event == save_setting:
        for row in save_file.rows:
            
            save_file.default_values[row] = window[row]
window.close()


